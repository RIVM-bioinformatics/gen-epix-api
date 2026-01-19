import re
from datetime import datetime
from pathlib import Path
from typing import Any, Literal

import polars as pl
import pytest
import xlsxwriter
from openpyxl import Workbook

# Initialize non-aggregated test data: tests incl. their result, scenarios, and the link between them
tests: list[dict[str, Any]] = []
scenario_ids: set[str] = set()
test_scenario_links: list[dict[str, str]] = []


def pytest_runtest_makereport(item: pytest.Item, call: pytest.CallInfo[Any]) -> None:
    """custom pytest hook to parse the docstring of the test function."""
    if call.when == "call":
        curr_scenarios = set()
        # Add scenario IDs from parent class
        parent = item.parent
        mark: pytest.Mark | None
        if isinstance(parent, pytest.Class):
            mark = parent.keywords.get("scenario_ids")
            if mark:
                curr_scenarios.update(re.split(r"\s*,\s*", mark.args[0]))
        # Add scenario IDs from test
        if isinstance(item, pytest.Function):
            mark = item.keywords.get("scenario_ids")
            if mark:
                curr_scenarios.update(re.split(r"\s*,\s*", mark.args[0]))
        else:
            raise AssertionError(f"Unknown item of type {type(item)}")
        # Add test
        tests.append(
            {
                "id": item.nodeid,
                "outcome": "PASS" if call.excinfo is None else "FAIL",
                "doc": item.function.__doc__,
                "datetime": datetime.now(),
                "duration": call.duration,
            }
        )
        # Add scenario
        scenario_ids.update(curr_scenarios)
        # Add scenario-test link
        test_scenario_links.extend(
            [
                {"scenario_id": scenario_id, "test_id": item.nodeid}
                for scenario_id in curr_scenarios
            ]
        )


def old_generate_excel_report(
    results: list[dict[str, str | list[str] | Literal["Pass", "Fail"] | float]],
    file_path: str,
) -> None:
    """
    Generate Excel file with two sheets:
    1. Aggregated by scenario ID
    2. Individual test results
    """
    aggregated_results: dict[str, dict[str, object]] = {}
    detailed_rows: list[tuple[str, str, str, str, float]] = []

    for result in results:
        test_name = str(result.get("test_name", ""))
        result_date = str(result.get("result_date", ""))
        test_result = str(result.get("result", "Fail"))
        duration = float(result.get("duration", 0.0))  # type: ignore[arg-type]
        scenario_ids: list[str] = result.get("scenario_ids") or []  # type: ignore[assignment]

        if not scenario_ids:
            detailed_rows.append((test_name, "", result_date, test_result, duration))

        for scenario_id in scenario_ids:
            detailed_rows.append(
                (test_name, scenario_id, result_date, test_result, duration)
            )
            # Initialize aggregation entry
            if scenario_id not in aggregated_results:
                aggregated_results[scenario_id] = {
                    "id": scenario_id,
                    "result_date": result_date,
                    "all_pass": True,
                }
            if test_result != "Pass":
                # flip to False on any Fail
                aggregated_results[scenario_id]["all_pass"] = False
            try:
                current_datetime = datetime.strptime(
                    str(aggregated_results[scenario_id]["result_date"]),
                    "%Y-%m-%d %H:%M:%S",
                )
                new_datetime = datetime.strptime(result_date, "%Y-%m-%d %H:%M:%S")
                if new_datetime > current_datetime:
                    aggregated_results[scenario_id]["result_date"] = result_date
            except ValueError:
                # overwrite if parsing fails
                aggregated_results[scenario_id]["result_date"] = result_date

    old_save_excel_report(file_path, aggregated_results, detailed_rows)


def old_save_excel_report(
    file_path: str,
    aggregated_results: dict[str, dict[str, object]],
    detailed_rows: list[tuple[str, str, str, str, float]],
) -> None:
    workbook = Workbook()
    aggregated_results_sheet = workbook.active
    assert aggregated_results_sheet is not None
    aggregated_results_sheet.title = "Aggregated"
    aggregated_results_sheet.append(["id", "result_date", "result"])
    for _, data in aggregated_results.items():
        aggregated_results_sheet.append(
            [data["id"], data["result_date"], "Pass" if data["all_pass"] else "Fail"]
        )
    detailed_results_sheet = workbook.create_sheet("Details")
    detailed_results_sheet.append(
        ["test_name", "scenario_id", "result_date", "result", "duration"]
    )
    for row in detailed_rows:
        detailed_results_sheet.append(list(row))

    out_path = Path(file_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)

    print(
        f"\n\nExcel report with linked test scenarios saved to: {out_path.resolve()}\n"
    )


def generate_excel_report(
    tests: list[dict[str, Any]], file_path: str, verbose: bool = True
) -> None:
    """
    Generate Excel file with two sheets:
    1. Individual test results
    2. Aggregated by scenario ID
    """
    test_df = pl.DataFrame(tests)
    scenario_df = pl.DataFrame([{"id": x} for x in sorted(scenario_ids)])
    test_scenario_link_df = pl.DataFrame(test_scenario_links)

    # Outer join test and test_scenario_link
    test_with_scenarios_df = test_df.join(
        test_scenario_link_df, left_on="id", right_on="test_id", how="left"
    ).rename({"id": "test_id"})

    # Aggregate data by scenario:
    # Count n_passed and n_failed separately
    # Add durations
    # Add status PASS if n_failed==0
    result_by_scenario_df = (
        # test_with_scenarios_df.filter(pl.col("scenario_id").is_not_null())
        test_with_scenarios_df.group_by("scenario_id")
        .agg(
            [
                pl.count().alias("n_total"),
                (pl.col("outcome") == "PASS").sum().alias("n_passed"),
                (pl.col("outcome") == "FAIL").sum().alias("n_failed"),
                pl.sum("duration").alias("duration"),
            ]
        )
        .with_columns(
            [
                pl.when(pl.col("n_failed") == 0)
                .then(pl.lit("PASS"))
                .otherwise(pl.lit("FAIL"))
                .alias("status")
            ]
        )
    )

    # Aggregate data by test:
    # Count n_scenarios
    # Add min(durations)
    # Add status PASS if n_failed==0 else FAIL
    result_by_test_df = test_with_scenarios_df.group_by("test_id").agg(
        [
            pl.col("scenario_id").drop_nulls().n_unique().alias("n_scenarios"),
            pl.min("duration").alias("min_duration"),
            pl.first("outcome").alias("status"),
            pl.first("doc").alias("doc"),
            pl.first("datetime").alias("datetime"),
        ]
    )

    # Write out
    with xlsxwriter.Workbook(file_path) as workbook:
        test_df.write_excel(workbook, worksheet="test")
        scenario_df.write_excel(workbook, worksheet="scenario")
        test_scenario_link_df.write_excel(workbook, worksheet="test_scenario_link")
        result_by_scenario_df.write_excel(workbook, worksheet="result_by_scenario")
        result_by_test_df.write_excel(workbook, worksheet="result_by_test")
    if verbose:
        print(f"Analysis complete. Results written to: {file_path}")
    return


def pytest_sessionfinish(session: pytest.Session, exitstatus: int) -> None:
    """custom pytest hook to perform actions at the end of the test session."""
    # only create report if there is data to report
    if tests and scenario_ids and test_scenario_links:
        generate_excel_report(tests, "test/output/test_report.xlsx")
